import io
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.models import Invoice, UploadLog
from app.schemas.invoice import InvoiceResponse, UploadLogResponse

router = APIRouter()


@router.get("/invoices", response_model=list[InvoiceResponse])
def list_invoices(
    vendor: str | None = Query(None, description="Filter by vendor name (partial match)"),
    payment_status: str | None = Query(None, description="Filter by payment_status"),
    db: Session = Depends(get_db),
):
    query = db.query(Invoice)
    if vendor:
        query = query.filter(Invoice.vendor_name.ilike(f"%{vendor}%"))
    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)
    return query.order_by(Invoice.upload_timestamp.desc()).all()


@router.get("/invoices/{invoice_id}", response_model=InvoiceResponse)
def get_invoice(invoice_id: int, db: Session = Depends(get_db)):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return invoice


@router.get("/export")
def export_invoices(db: Session = Depends(get_db)):
    invoices = db.query(Invoice).order_by(Invoice.upload_timestamp.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "ID", "Invoice Number", "Vendor Name", "Invoice Date",
        "Amount", "Tax Amount", "Total Amount", "Payment Status",
        "Source File", "Upload Timestamp", "Processing Status",
    ]

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, inv in enumerate(invoices, start=2):
        ws.cell(row=row_idx, column=1, value=inv.id)
        ws.cell(row=row_idx, column=2, value=inv.invoice_number)
        ws.cell(row=row_idx, column=3, value=inv.vendor_name)
        ws.cell(row=row_idx, column=4, value=str(inv.invoice_date) if inv.invoice_date else None)
        ws.cell(row=row_idx, column=5, value=float(inv.amount) if inv.amount is not None else None)
        ws.cell(row=row_idx, column=6, value=float(inv.tax_amount) if inv.tax_amount is not None else None)
        ws.cell(row=row_idx, column=7, value=float(inv.total_amount) if inv.total_amount is not None else None)
        ws.cell(row=row_idx, column=8, value=inv.payment_status)
        ws.cell(row=row_idx, column=9, value=inv.source_file)
        ws.cell(row=row_idx, column=10, value=str(inv.upload_timestamp) if inv.upload_timestamp else None)
        ws.cell(row=row_idx, column=11, value=inv.processing_status)

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices.xlsx"},
    )


@router.get("/logs", response_model=list[UploadLogResponse])
def list_logs(db: Session = Depends(get_db)):
    return db.query(UploadLog).order_by(UploadLog.uploaded_at.desc()).all()


@router.get("/health")
def health_check():
    return {"status": "ok"}
